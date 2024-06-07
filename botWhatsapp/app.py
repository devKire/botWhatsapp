import openpyxl
from urllib.parse import quote
import webbrowser
from time import sleep
import pyautogui
import os
import sys
import re
import threading
import tkinter as tk
from tkinter import messagebox

running = False  # Variável global para controlar o estado de execução

# Variáveis globais para os tempos de espera
tempo_carregamento = 20
tempo_envio = 15
tempo_espera_envio = 5

def enviar_mensagens():
    global running, tempo_carregamento, tempo_envio, tempo_espera_envio

    # Abrir WhatsApp Web e aguardar o carregamento
    webbrowser.open('https://web.whatsapp.com/')
    sleep(tempo_carregamento)  # Aguardar o tempo suficiente para o carregamento

    # Ler planilha e guardar informações sobre nome, telefone e data de vencimento
    try:
        workbook = openpyxl.load_workbook('numeros.xlsx')
        pagina_numeros = workbook['Sheet1']
    except FileNotFoundError:
        messagebox.showerror("Erro", "Arquivo 'numeros.xlsx' não encontrado.")
        return
    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao abrir o arquivo 'numeros.xlsx': {e}")
        return

    # Carregar ou criar o arquivo para rastrear números enviados
    try:
        workbook_enviados = openpyxl.load_workbook('enviados.xlsx')
        planilha_enviados = workbook_enviados.active
    except FileNotFoundError:
        workbook_enviados = openpyxl.Workbook()
        planilha_enviados = workbook_enviados.active
        planilha_enviados.append(['Número'])
        workbook_enviados.save('enviados.xlsx')

    mensagem = text_mensagem.get("1.0", tk.END).strip()

    for linha in pagina_numeros.iter_rows(min_row=2):
        if not running:
            break  # Pausar se running for False
        num = linha[0].value

        # Verificar se o número já foi enviado
        if any(num == row[0].value for row in planilha_enviados.iter_rows()):
            continue  # Pular para o próximo número se já foi enviado

        # Criar link personalizado do WhatsApp e enviar mensagem para cada cliente
        try:
            link_mensagem_whatsapp = f'https://web.whatsapp.com/send?phone={num}&text={quote(mensagem)}'
            webbrowser.open(link_mensagem_whatsapp)
            sleep(tempo_envio)  # Aguardar a página carregar
            
            # Pressionar Enter para enviar a mensagem
            pyautogui.press('enter')
            sleep(tempo_espera_envio)  # Aguardar o envio da mensagem
            
            # Fechar a aba do navegador
            pyautogui.hotkey('ctrl', 'w')
            sleep(2)

            # Registrar o número como enviado
            planilha_enviados.append([num])
            workbook_enviados.save('enviados.xlsx')
        except Exception as e:
            messagebox.showerror("Erro", f"Não foi possível enviar mensagem para {num}: {e}")
            with open('erros.csv', 'a', newline='', encoding='utf-8') as arquivo:
                arquivo.write(f'{num}{os.linesep}')
    
    # Encerrar o script ao final do loop
    if running:
        print("Envio de mensagens concluído. Encerrando o script.")
        messagebox.showinfo("Sucesso", "Envio de mensagens concluído.")
        running = False
def start():
    global running
    if not running:
        running = True
        threading.Thread(target=enviar_mensagens).start()

def pause():
    global running
    running = False

def adicionar_numeros():
    numeros = text_numeros.get("1.0", tk.END).strip()
    numeros_validos = re.findall(r'\+\d{1,3}\s?\(?\d{2,4}\)?\s?\d{3,5}[-\s]?\d{4,5}', numeros)
    
    if numeros_validos:
        try:
            workbook = openpyxl.load_workbook('numeros.xlsx')
            sheet = workbook.active
            for numero in numeros_validos:
                sheet.append([numero])
            workbook.save('numeros.xlsx')
            text_numeros.delete("1.0", tk.END)
            messagebox.showinfo("Sucesso", f"Números adicionados com sucesso.")
            atualizar_listbox_numeros()  # Atualizar a lista após adição bem-sucedida
        except FileNotFoundError:
            messagebox.showerror("Erro", "Arquivo 'numeros.xlsx' não encontrado.")
        except Exception as e:
            messagebox.showerror("Erro", f"Não foi possível adicionar os números: {e}")
    else:
        messagebox.showwarning("Atenção", "Por favor, insira números válidos e não repetidos no formato correto para números de WhatsApp.")

def remover_numeros():
    numeros = listbox_numeros.curselection()
    if numeros:
        try:
            workbook = openpyxl.load_workbook('numeros.xlsx')
            sheet = workbook.active
            numeros = sorted(numeros, reverse=True)  # Remover em ordem reversa para evitar índices inválidos
            for numero_index in numeros:
                sheet.delete_rows(numero_index + 2)
            workbook.save('numeros.xlsx')
            atualizar_listbox_numeros()
            messagebox.showinfo("Sucesso", f"Números removidos com sucesso.")
        except FileNotFoundError:
            messagebox.showerror("Erro", "Arquivo 'numeros.xlsx' não encontrado.")
        except Exception as e:
            messagebox.showerror("Erro", f"Não foi possível remover os números: {e}")
    else:
        messagebox.showwarning("Atenção", "Por favor, selecione pelo menos um número.")

def remover_todos():
    if messagebox.askokcancel("Confirmação", "Tem certeza de que deseja remover todos os números?"):
        try:
            workbook = openpyxl.load_workbook('numeros.xlsx')
            sheet = workbook.active
            last_row = sheet.max_row
            
            # Remover todas as linhas exceto a primeira (cabeçalho)
            for _ in range(last_row - 1):
                sheet.delete_rows(2)
            
            workbook.save('numeros.xlsx')
            atualizar_listbox_numeros()
            messagebox.showinfo("Sucesso", "Todos os números removidos com sucesso.")
        except FileNotFoundError:
            messagebox.showerror("Erro", "Arquivo 'numeros.xlsx' não encontrado.")
        except Exception as e:
            messagebox.showerror("Erro", f"Não foi possível remover os números: {e}")



def atualizar_listbox_numeros():
    listbox_numeros.delete(0, tk.END)
    try:
        workbook = openpyxl.load_workbook('numeros.xlsx')
        sheet = workbook.active
        for row in sheet.iter_rows(min_row=2, values_only=True):
            listbox_numeros.insert(tk.END, row[0])
    except FileNotFoundError:
        messagebox.showwarning("Aviso", "Arquivo 'numeros.xlsx' não encontrado. Por favor, adicione números para enviar mensagens.")
    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao carregar os números: {e}")

def atualizar_tempos():
    global tempo_carregamento, tempo_envio, tempo_espera_envio
    try:
        tempo_carregamento = int(entry_tempo_carregamento.get())
        tempo_envio = int(entry_tempo_envio.get())
        tempo_espera_envio = int(entry_tempo_espera_envio.get())
        messagebox.showinfo("Sucesso", "Tempos atualizados com sucesso.")
    except ValueError:
        messagebox.showerror("Erro", "Por favor, insira valores válidos para os tempos.")

# Criar a janela principal
root = tk.Tk()
root.title("WhatsApp Message Sender")

# Criar um frame principal
main_frame = tk.Frame(root)
main_frame.pack(fill=tk.BOTH, expand=1)

# Criar um canvas
canvas = tk.Canvas(main_frame)
canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=1)

# Adicionar uma barra de rolagem ao canvas
scrollbar = tk.Scrollbar(main_frame, orient=tk.VERTICAL, command=canvas.yview)
scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

# Configurar o canvas para usar a barra de rolagem
canvas.configure(yscrollcommand=scrollbar.set)
canvas.bind('<Configure>', lambda e: canvas.configure(scrollregion=canvas.bbox("all")))

# Criar um frame secundário dentro do canvas com uma borda
second_frame = tk.Frame(canvas, bd=2, relief=tk.RIDGE)

# Adicionar o frame secundário a uma janela no canvas
canvas.create_window((0,0), window=second_frame, anchor="nw", tags="frame")

def resize(event):
    canvas.configure(scrollregion=canvas.bbox("all"))

second_frame.bind("<Configure>", resize)

def on_canvas_configure(event):
    canvas.itemconfig("frame", width=event.width)

canvas.bind("<Configure>", on_canvas_configure)

# Adicionar botões Play e Pause com um estilo de botão
play_button = tk.Button(second_frame, text="Play", command=start, relief=tk.RAISED, bg="#4CAF50", fg="white", padx=10, pady=5)
play_button.pack(pady=10, fill=tk.X)

pause_button = tk.Button(second_frame, text="Pause", command=pause, relief=tk.RAISED, bg="#f44336", fg="white", padx=10, pady=5)
pause_button.pack(pady=10, fill=tk.X)

# Adicionar campos para inserir novos números com uma borda e cor de fundo
label_numeros = tk.Label(second_frame, text="Adicionar Números neste formato: \n +XX XX XXXX-XXXX, +XX XXXX-XXXX, \n +XX (XX) XXXX-XXXX, +XXXXXXXXXXX, ...", bg="#f0f0f0", padx=10, pady=5)
label_numeros.pack(pady=5, fill=tk.X)
text_numeros = tk.Text(second_frame, height=10, bd=2, relief=tk.SOLID)
text_numeros.pack(pady=5, fill=tk.X)
add_button = tk.Button(second_frame, text="Adicionar", command=adicionar_numeros, relief=tk.RAISED, bg="#2196F3", fg="white", padx=10, pady=5)
add_button.pack(pady=5, fill=tk.X)

# Adicionar campo para selecionar números para remover com uma borda e cor de fundo
label_remover_numeros = tk.Label(second_frame, text="Selecionar Números para Remover:", bg="#f0f0f0", padx=10, pady=5)
label_remover_numeros.pack(pady=5, fill=tk.X)
scrollbar_numeros = tk.Scrollbar(second_frame)
scrollbar_numeros.pack(side=tk.RIGHT, fill=tk.Y)

listbox_numeros = tk.Listbox(second_frame, yscrollcommand=scrollbar_numeros.set, selectmode=tk.MULTIPLE, bd=2, relief=tk.SOLID)
listbox_numeros.pack(pady=5, fill=tk.X)

scrollbar_numeros.config(command=listbox_numeros.yview)

remove_button = tk.Button(second_frame, text="Remover Selecionados", command=remover_numeros, relief=tk.RAISED, bg="#f44336", fg="white", padx=10, pady=5)
remove_button.pack(pady=5, fill=tk.X)
remove_button = tk.Button(second_frame, text="Remover Todos", command=remover_todos, relief=tk.RAISED, bg="#f44336", fg="white", padx=10, pady=5)
remove_button.pack(pady=5, fill=tk.X)

# Adicionar campo para editar a mensagem com uma borda e cor de fundo
label_mensagem = tk.Label(second_frame, text="Mensagem a ser enviada:", bg="#f0f0f0", padx=10, pady=5)
label_mensagem.pack(pady=5, fill=tk.X)
text_mensagem = tk.Text(second_frame, height=10, bd=2, relief=tk.SOLID)
text_mensagem.pack(pady=5, fill=tk.X)
text_mensagem.insert(tk.END, "Opa mano, tudo certo? Criei um grupo do Whatsapp focado em alguns jogos, se tiver interesse é só entrar lá😎👍\n📱 Grupo Geral do WhatsApp: https://chat.whatsapp.com/HggfZseAPPvFwq3FeP5vxz\n🔗 Link do Discord: https://discord.com/invite/gjZhfmfSKH")

# Adicionar campos para editar os tempos de espera com uma borda e cor de fundo
label_tempos = tk.Label(second_frame, text="Editar Tempos (segundos):", bg="#f0f0f0", padx=10, pady=5)
label_tempos.pack(pady=5, fill=tk.X)

label_tempo_carregamento = tk.Label(second_frame, text="Tempo de Carregamento:", bg="#f0f0f0", padx=10, pady=5)
label_tempo_carregamento.pack(pady=5, fill=tk.X)
entry_tempo_carregamento = tk.Entry(second_frame, bd=2, relief=tk.SOLID)
entry_tempo_carregamento.insert(0, str(tempo_carregamento))
entry_tempo_carregamento.pack(pady=5, fill=tk.X)

label_tempo_envio = tk.Label(second_frame, text="Tempo de Envio:", bg="#f0f0f0", padx=10, pady=5)
label_tempo_envio.pack(pady=5, fill=tk.X)
entry_tempo_envio = tk.Entry(second_frame, bd=2, relief=tk.SOLID)
entry_tempo_envio.insert(0, str(tempo_envio))
entry_tempo_envio.pack(pady=5, fill=tk.X)

label_tempo_espera_envio = tk.Label(second_frame, text="Tempo de Espera após Envio:", bg="#f0f0f0", padx=10, pady=5)
label_tempo_espera_envio.pack(pady=5, fill=tk.X)
entry_tempo_espera_envio = tk.Entry(second_frame, bd=2, relief=tk.SOLID)
entry_tempo_espera_envio.insert(0, str(tempo_espera_envio))
entry_tempo_espera_envio.pack(pady=5, fill=tk.X)

button_atualizar_tempos = tk.Button(second_frame, text="Atualizar Tempos", command=atualizar_tempos, relief=tk.RAISED, bg="#f44336", fg="white", padx=10, pady=5)
button_atualizar_tempos.pack(pady=10, fill=tk.X)


# Atualizar a lista de números inicialmente
atualizar_listbox_numeros()

# Rodar a aplicação
root.mainloop()
